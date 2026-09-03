#!/usr/bin/env python3
"""Processador FM26 Moneyball com saída XLSX limpa e compatível com Excel Mobile."""

from __future__ import annotations

import argparse
import math
import re
import shutil
import sqlite3
import sys
import tempfile
import zipfile
from bisect import bisect_left, bisect_right
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Constantes removidas; agora calculadas dinamicamente via get_date_info()

COLORS = {
    "title": "0B3D2E",
    "header": "0E4A36",
    "bad": "F4CCCC",
    "medium": "FFF2CC",
    "good": "D9EAD3",
    "excellent": "CFE2F3",
    "neutral": "FFFFFF",
    "subtitle": "EAF3EE",
}

# Papéis de linha (sem GK) — cada um vira coluna própria
OUTFIELD_ROLES = ["CD", "LB", "RB", "CDM", "CM", "CAM", "LM", "RM", "LW", "RW", "ST"]

PLAYERS_HEADERS = [
    "Rank", "Jogador", "Posição",
    "Role CD", "Role LB", "Role RB", "Role CDM", "Role CM", "Role CAM", "Role LM", "Role RM", "Role LW", "Role RW", "Role ST",
    "Idade", "Time", "Liga", "Divisão", "País",
    "Overall do Footlord (0-100)", "Overall início da temporada", "Variação overall na temporada", "Overall início da carreira", "Variação overall desde início da carreira",
    "Jogos no ano", "Cartões amarelos", "Cartões vermelhos", "Cartões totais", "Amarelos/partida", "Vermelhos/partida", 
    "Gols no ano", "Gols sem pênalti", "Pênaltis batidos", "Gols de pênalti", "Pênaltis perdidos (total)", "Pênaltis perdidos (fora)", "Taxa conversão pênalti", "xGP (pênaltis esperados)", "Gols pênalti vs xGP",
    "Assistências no ano", "G+A no ano", "Gols/partida", "Assistências/partida", "G+A/partida", "Média de avaliação", "Minutos estimados",
    "Minutos/partida", "Minutos/gol", "Minutos/assistência", "MVP do time", "Pior nota do time",
    "MVP %", "Pior nota %", "Clean sheets", "Valor de mercado em €", "Salário em €",
    "Contrato (anos)", "Status", "Empréstimo", "Lista de transferências", "PAC", "SHO", "PAS", "DEF", "PHY", "MEN", "GKP", "ID interno",
]

GK_HEADERS = [
    "Rank", "Goleiro", "Time", "Liga", "Partidas GK", "Minutos", "Clean sheets", "Clean sheets/partida", "xG adversário", "Gols sofridos", "Gols/partida GK",
    "Gols evitados vs xG", "Gols evitados/partida", "Gols evitados/90", "Taxa de prevenção", "Gols/xG", 
    "Pênaltis enfrentados", "Pênaltis defendidos", "Taxa defesa pênalti", "xGP (pênaltis esperados)", "Pênaltis defendidos vs xGP",
    "Classificação", "Elegível (mín. 5 jogos)", "Overall do Footlord", "ID interno",
]

TIMES_HEADERS = [
    "Rank overall", "Time", "Liga", "Competição continental", "Jogadores no time", "Média de idade",
    "Partidas no ano", "Vitórias no ano", "Empates no ano", "Derrotas no ano", "Cartões amarelos no ano", "Cartões vermelhos no ano", "Cartões totais no ano", "Amarelos/partida", "Vermelhos/partida", "Partidas antes 1/jan", "Vitórias antes", "Empates antes", "Derrotas antes",
    "Partidas depois 1/jan", "Vitórias depois", "Empates depois", "Derrotas depois", "Maior sequência vitórias antes", "Maior sequência derrotas antes",
    "Maior sequência vitórias depois", "Maior sequência derrotas depois", "Gols marcados no ano", "Gols sofridos no ano", "Gols marcados/partida",
    "Gols sofridos/partida", "Chutes totais", "Chutes/partida", "Chutes no alvo", "% no alvo", "Conversão de chutes", "Gols marcados antes", "Gols sofridos antes", "Gols marcados/partida antes", "Gols sofridos/partida antes",
    "Gols marcados depois", "Gols sofridos depois", "Gols marcados/partida depois", "Gols sofridos/partida depois", "Jogos sem marcar",
    "Jogos sem sofrer gols", "Sem marcar antes", "Sem sofrer gols antes", "Sem marcar depois", "Sem sofrer gols depois", "Maior placar aplicado no ano",
    "Para quem — maior placar ano", "Pior placar sofrido no ano", "Para quem — pior placar ano", "Maior placar aplicado antes", "Para quem — maior placar antes",
    "Pior placar sofrido antes", "Para quem — pior placar antes", "Maior placar aplicado depois", "Para quem — maior placar depois", "Pior placar sofrido depois",
    "Para quem — pior placar depois", "Overall início do ano", "Overall fechamento agosto", "Overall final do ano", "Expectativa — posição por overall inicial",
    "Posição na liga no save", "Diferença expectativa/posição", "Média overall adversário — vitória", "Média overall adversário — empate",
    "Média overall adversário — derrota", "Valor do time — elenco em €", "Média salário dos jogadores em €", "Saldo bancário em €",
    "Gastos janela janeiro em €", "Vendas janela janeiro em €", "Overall médio compras janeiro", "Overall médio vendas janeiro",
    "Gastos janela agosto em €", "Vendas janela agosto em €", "Overall médio compras agosto", "Overall médio vendas agosto",
]

HEADER_TRANSLATIONS_EN = {
    "Rank": "Rank", "Quality Index (0-100)": "Quality Index (0-100)", "Quality Index": "Quality Index", "Score Moneyball (0-100)": "Moneyball Score (0-100)", "Score Moneyball": "Moneyball Score", "MVP %": "MVP %", "Clean sheets": "Clean sheets", "Clean sheets/partida": "Clean sheets/match", "Status": "Status", "PAC": "PAC", "SHO": "SHO", "PAS": "PAS", "DEF": "DEF", "PHY": "PHY", "MEN": "MEN", "GKP": "GKP", "Gols/xG": "Goals/xG",
    "Jogador": "Player", "Posição": "Position", "Role CD": "Role CD", "Role LB": "Role LB", "Role RB": "Role RB", "Role CDM": "Role CDM", "Role CM": "Role CM", "Role CAM": "Role CAM", "Role LM": "Role LM", "Role RM": "Role RM", "Role LW": "Role LW", "Role RW": "Role RW", "Role ST": "Role ST", "Overall do Footlord (0-100)": "Footlord Overall (0-100)", "Overall do Footlord": "Footlord Overall", "Idade": "Age", "Time": "Team", "Liga": "League", "Divisão": "Division", "País": "Country",
    "Jogos no ano": "Matches in year", "Gols no ano": "Goals in year", "Assistências no ano": "Assists in year", "G+A no ano": "G+A in year", "Gols/partida": "Goals/match", "Assistências/partida": "Assists/match", "G+A/partida": "G+A/match", "Média de avaliação": "Average rating", "Minutos estimados": "Estimated minutes", "Partidas com minutos": "Matches with minutes", "Minutos/partida": "Minutes/match", "Minutos/gol": "Minutes/goal", "Minutos/assistência": "Minutes/assist", "MVP do time": "Team MVP", "Pior nota do time": "Team worst rating", "Partidas com rating": "Matches with rating", "Pior nota %": "Worst rating %", "Partidas GK": "GK matches", "xG adversário GK": "Opponent xG GK", "Gols sofridos GK": "Goals conceded GK", "Gols evitados vs xG": "Goals prevented vs xG", "xG/partida GK": "xG/match GK", "Gols/partida GK": "Goals/match GK", "Gols/xG GK": "Goals/xG GK", "Valor de mercado em €": "Market value in €", "Salário em €": "Salary in €", "Contrato (anos)": "Contract (years)", "Empréstimo": "Loan", "Lista de transferências": "Transfer listed", "ID interno": "Internal ID",
    "Goleiro": "Goalkeeper", "Minutos": "Minutes", "xG adversário": "Opponent xG", "Gols sofridos": "Goals conceded", "Gols evitados/partida": "Goals prevented/match", "Gols evitados/90": "Goals prevented/90", "Taxa de prevenção": "Prevention rate", "Classificação": "Rating", "Elegível (mín. 5 jogos)": "Eligible (min. 5 matches)",
    "Rank cartões": "Cards rank", "Cartões amarelos": "Yellow cards", "Cartões vermelhos": "Red cards", "Cartões totais": "Total cards", "Amarelos/partida": "Yellow cards/match", "Vermelhos/partida": "Red cards/match", "Cartões amarelos no ano": "Yellow cards in year", "Cartões vermelhos no ano": "Red cards in year", "Cartões totais no ano": "Total cards in year", "Overall início da temporada": "Overall at season start", "Overall atual": "Current overall", "Variação overall na temporada": "Overall change in season", "Overall início da carreira": "Overall at career start", "Variação overall desde início da carreira": "Overall change since career start",
    "Rank overall": "Overall rank", "Competição continental": "Continental competition", "Jogadores no time": "Players in team", "Média de idade": "Average age", "Partidas no ano": "Matches in year", "Vitórias no ano": "Wins in year", "Empates no ano": "Draws in year", "Derrotas no ano": "Losses in year", "Partidas antes 1/jan": "Matches before Jan 1", "Vitórias antes": "Wins before", "Empates antes": "Draws before", "Derrotas antes": "Losses before", "Partidas depois 1/jan": "Matches after Jan 1", "Vitórias depois": "Wins after", "Empates depois": "Draws after", "Derrotas depois": "Losses after", "Maior sequência vitórias antes": "Longest win streak before", "Maior sequência derrotas antes": "Longest loss streak before", "Maior sequência vitórias depois": "Longest win streak after", "Maior sequência derrotas depois": "Longest loss streak after", "Gols marcados no ano": "Goals scored in year", "Gols sofridos no ano": "Goals conceded in year", "Gols marcados/partida": "Goals scored/match", "Gols sofridos/partida": "Goals conceded/match", "Chutes totais": "Total shots", "Chutes/partida": "Shots/match", "Chutes no alvo": "Shots on target", "% no alvo": "% on target", "Conversão de chutes": "Shot conversion", "Gols marcados antes": "Goals scored before", "Gols sofridos antes": "Goals conceded before", "Gols marcados/partida antes": "Goals scored/match before", "Gols sofridos/partida antes": "Goals conceded/match before", "Gols marcados depois": "Goals scored after", "Gols sofridos depois": "Goals conceded after", "Gols marcados/partida depois": "Goals scored/match after", "Gols sofridos/partida depois": "Goals conceded/match after", "Jogos sem marcar": "Matches without scoring", "Jogos sem sofrer gols": "Clean sheets", "Sem marcar antes": "Without scoring before", "Sem sofrer gols antes": "Clean sheets before", "Sem marcar depois": "Without scoring after", "Sem sofrer gols depois": "Clean sheets after", "Maior placar aplicado no ano": "Biggest win in year", "Para quem — maior placar ano": "Opponent — biggest win in year", "Pior placar sofrido no ano": "Worst loss in year", "Para quem — pior placar ano": "Opponent — worst loss in year", "Maior placar aplicado antes": "Biggest win before", "Para quem — maior placar antes": "Opponent — biggest win before", "Pior placar sofrido antes": "Worst loss before", "Para quem — pior placar antes": "Opponent — worst loss before", "Maior placar aplicado depois": "Biggest win after", "Para quem — maior placar depois": "Opponent — biggest win after", "Pior placar sofrido depois": "Worst loss after", "Para quem — pior placar depois": "Opponent — worst loss after", "Overall início do ano": "Overall at start of year", "Overall fechamento agosto": "Overall at August deadline", "Overall final do ano": "Overall at end of year", "Expectativa — posição por overall inicial": "Expected position by starting overall", "Posição na liga no save": "League position in save", "Diferença expectativa/posição": "Expectation/position difference", "Média overall adversário — vitória": "Average opponent overall — win", "Média overall adversário — empate": "Average opponent overall — draw", "Média overall adversário — derrota": "Average opponent overall — loss", "Valor do time — elenco em €": "Team value — squad in €", "Média salário dos jogadores em €": "Average player salary in €", "Saldo bancário em €": "Bank balance in €", "Gastos janela janeiro em €": "January window spending in €", "Vendas janela janeiro em €": "January window sales in €", "Overall médio compras janeiro": "Average overall — January signings", "Overall médio vendas janeiro": "Average overall — January sales", "Gastos janela agosto em €": "August window spending in €", "Vendas janela agosto em €": "August window sales in €", "Overall médio compras agosto": "Average overall — August signings", "Overall médio vendas agosto": "Average overall — August sales",
    "Gols sem pênalti": "Non-penalty goals", "Pênaltis batidos": "Penalties taken", "Gols de pênalti": "Penalty goals", "Pênaltis perdidos (total)": "Penalties missed (total)", "Pênaltis perdidos (fora)": "Penalties missed (out)", "Taxa conversão pênalti": "Penalty conversion rate", "xGP (pênaltis esperados)": "xGP (expected penalties)", "Gols pênalti vs xGP": "Penalty goals vs xGP",
    "Pênaltis enfrentados": "Penalties faced", "Pênaltis defendidos": "Penalties saved", "Taxa defesa pênalti": "Penalty save rate", "Pênaltis defendidos vs xGP": "Penalties saved vs xGP",
}


def localized_headers(headers, language):
    return [HEADER_TRANSLATIONS_EN.get(header, header) for header in headers] if language == "en" else headers

NUM_FORMAT_INT = '#,##0;[Red]-#,##0'
NUM_FORMAT_DEC1 = '0.0;[Red]-0.0'
NUM_FORMAT_DEC2 = '0.00;[Red]-0.00'
NUM_FORMAT_DEC3 = '0.000;[Red]-0.000'
NUM_FORMAT_PCT = '0.0%;[Red]-0.0%'

THIN_BLACK = Side(style="thin", color="000000")
BORDER = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
TITLE_FONT = Font(name="Aptos Display", size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Aptos", size=9, color="111827")
FILLS = {name: PatternFill("solid", fgColor=value) for name, value in COLORS.items()}


def emit(percent: int, message: str) -> None:
    print(f"PROGRESS|{percent}|{message}", flush=True)


def sf(value, default=0.0):
    try:
        return default if value is None or value == "" else float(value)
    except (TypeError, ValueError):
        return default


def si(value, default=0):
    try:
        return default if value is None or value == "" else int(float(value))
    except (TypeError, ValueError):
        return default


def clip(value, low=0.0, high=100.0):
    return max(low, min(high, sf(value)))


def parse_ids(raw):
    return [si(value) for value in str(raw or "").split("_") if str(value).strip()]


def parse_values(raw):
    values = []
    for value in str(raw or "").split("_"):
        try:
            values.append(float(value))
        except (TypeError, ValueError):
            pass
    return values


def parse_count_blocks(raw):
    total = 0.0
    for item in str(raw or "").split(","):
        if ":" not in item:
            continue
        try:
            total += float(item.rsplit(":", 1)[1])
        except (TypeError, ValueError):
            pass
    return total


def parse_rating_blocks(raw):
    values = []
    for item in str(raw or "").split(","):
        if ":" not in item:
            continue
        for value in item.rsplit(":", 1)[1].split("|"):
            try:
                values.append(float(value))
            except (TypeError, ValueError):
                pass
    return values


def parse_match_ratings(raw):
    result = {}
    for item in str(raw or "").split(";"):
        if "|" not in item:
            continue
        player_id, rating = item.split("|", 1)
        try:
            result[int(player_id)] = float(rating)
        except (TypeError, ValueError):
            pass
    return result


def id_set(value):
    return {int(part) for part in str(value or "").split("_") if part.isdigit()}


def rating_ids(value):
    ids = set()
    for part in str(value or "").split(";"):
        if "|" in part:
            try:
                ids.add(int(part.split("|", 1)[0]))
            except (TypeError, ValueError):
                pass
    return ids


def parse_role_levels(raw):
    """Parse 'GK:10.00,CD:1.78,...' -> dict role -> float."""
    result = {}
    for item in str(raw or "").split(","):
        if ":" not in item:
            continue
        role, value = item.split(":", 1)
        try:
            result[role.strip()] = float(value)
        except (TypeError, ValueError):
            pass
    return result


def parse_events(raw):
    """Eventos de partida. Minuto pode passar de 119 (pênaltis de desempate)."""
    events = []
    for item in str(raw or "").split(";"):
        parts = item.split("|")
        if len(parts) != 3:
            continue
        try:
            # Não cortar em 120: desempates usam minuto >= 120
            events.append((max(0, int(parts[0])), int(parts[1]), int(parts[2])))
        except (TypeError, ValueError):
            pass
    return events


def estimate_minutes(lineup, entered, subbed, events):
    minutes = {pid: 90.0 for pid in lineup}
    for pid in entered:
        minutes[pid] = 45.0
    for minute, code, pid in events:
        if code == 6 and pid in minutes:
            minutes[pid] = float(minute)
        elif code == 5:
            minutes[pid] = 90.0 - float(minute)
    return minutes


def percentile(values, value):
    if not values:
        return 50.0
    sorted_values = sorted(values)
    idx = bisect_left(sorted_values, value)
    return clip(idx / len(sorted_values) * 100)


def get_date_info(cur):
    gen = dict(cur.execute("SELECT * FROM general LIMIT 1").fetchone())
    current_day = si(gen["date"])
    starting_year = si(gen.get("starting_year"), 2025)
    # 1º de julho do ano inicial é o dia 0
    start_date = datetime(starting_year, 7, 1)
    current_date = start_date + timedelta(days=current_day)
    
    # 1º de janeiro do ano civil ATUAL (em relação ao current_day)
    jan_1st = datetime(current_date.year, 1, 1)
    # Se hoje é antes de julho, o 1º de janeiro relevante é o deste ano.
    # Se hoje é julho ou depois, o 1º de janeiro relevante é o do ano que vem?
    # O usuário quer dividir a temporada. Temporada europeia: Jul-Jun.
    # Se estamos em Fev/2030, o "antes de 1/jan" é Jul/2029-Dez/2029.
    # Se estamos em Out/2030, o "antes de 1/jan" é Jul/2030-Dez/2030 (vazio) e "depois" é Jan/2031+ (vazio).
    # Regra: O 1º de janeiro que divide a temporada atual.
    # Se mês >= 7: o 1º de jan é o do ano seguinte.
    # Se mês < 7: o 1º de jan é o deste ano.
    target_jan = datetime(current_date.year if current_date.month < 7 else current_date.year + 1, 1, 1)
    jan_cutoff = (target_jan - start_date).days
    
    return {
        "current_day": current_day,
        "current_date": current_date,
        "start_date": start_date,
        "jan_cutoff": jan_cutoff,
        "managed_id": si(gen.get("managed_team_id"), -1),
        "managed_name": gen.get("managed_team_name") or "Clube"
    }

def build_team_rows(cur, date_info, players, played_matches, team_cards):
    teams = {row["id"]: dict(row) for row in cur.execute("SELECT * FROM teams2").fetchall()}
    leagues = {row["id"]: dict(row) for row in cur.execute("SELECT * FROM leagues2").fetchall()}
    for team_id, team in teams.items():
        league_id = si(team.get("league_id"), -1)
        team.update({
            "league": leagues.get(league_id, {}).get("name", "Desconhecida"),
            "all": {"games": 0, "wins": 0, "draws": 0, "losses": 0, "goals_for": 0, "goals_against": 0, "shots": 0, "on_target": 0, "clean_sheets": 0, "failed_to_score": 0, "streak_w": 0, "streak_l": 0, "cur_w": 0, "cur_l": 0, "opp_ov_w": [], "opp_ov_d": [], "opp_ov_l": [], "biggest_win": (0, ""), "worst_loss": (0, "")},
            "before": {"games": 0, "wins": 0, "draws": 0, "losses": 0, "goals_for": 0, "goals_against": 0, "clean_sheets": 0, "failed_to_score": 0, "streak_w": 0, "streak_l": 0, "cur_w": 0, "cur_l": 0, "biggest_win": (0, ""), "worst_loss": (0, "")},
            "after": {"games": 0, "wins": 0, "draws": 0, "losses": 0, "goals_for": 0, "goals_against": 0, "clean_sheets": 0, "failed_to_score": 0, "streak_w": 0, "streak_l": 0, "cur_w": 0, "cur_l": 0, "biggest_win": (0, ""), "worst_loss": (0, "")},
            "players": [p for p in players.values() if p["team_id"] == team_id],
            "yellow_cards": team_cards.get(team_id, {}).get("yellow", 0),
            "red_cards": team_cards.get(team_id, {}).get("red", 0),
        })
    for raw in played_matches:
        match = dict(raw)
        t1, t2 = si(match["team_1_id"]), si(match["team_2_id"])
        g1, g2 = si(match["goals_1"]), si(match["goals_2"])
        date = si(match["date"])
        s1 = si(match["shots_on_1"]) + si(match["shots_off_1"])
        s2 = si(match["shots_on_2"]) + si(match["shots_off_2"])
        ot1, ot2 = si(match["shots_on_1"]), si(match["shots_on_2"])
        for tid, side in [(t1, 1), (t2, 2)]:
            if tid not in teams: continue
            t, other = teams[tid], 3 - side
            og, os, oot = si(match[f"goals_{other}"]), si(match[f"shots_on_{other}"]) + si(match[f"shots_off_{other}"]), si(match[f"shots_on_{other}"])
            my_g, my_s, my_ot = si(match[f"goals_{side}"]), si(match[f"shots_on_{side}"]) + si(match[f"shots_off_{side}"]), si(match[f"shots_on_{side}"])
            period = t["before"] if date < date_info["jan_cutoff"] else t["after"]
            for d in [t["all"], period]:
                d["games"] += 1
                d["goals_for"] += my_g
                d["goals_against"] += og
                d["clean_sheets"] += og == 0
                d["failed_to_score"] += my_g == 0
                if d == t["all"]:
                    d["shots"] += my_s
                    d["on_target"] += my_ot
                if my_g > og:
                    d["wins"] += 1
                    d["cur_w"] += 1
                    d["cur_l"] = 0
                    d["streak_w"] = max(d["streak_w"], d["cur_w"])
                    if d == t["all"]:
                        opp_ov = teams.get(si(match[f"team_{other}_id"]), {}).get("temp_value", 50.0)
                        d["opp_ov_w"].append(opp_ov)
                    if my_g - og > d["biggest_win"][0]: d["biggest_win"] = (my_g - og, teams.get(si(match[f"team_{other}_id"]), {}).get("name", "Desconhecido"))
                elif my_g == og:
                    d["draws"] += 1
                    d["cur_w"] = d["cur_l"] = 0
                    if d == t["all"]:
                        opp_ov = teams.get(si(match[f"team_{other}_id"]), {}).get("temp_value", 50.0)
                        d["opp_ov_d"].append(opp_ov)
                else:
                    d["losses"] += 1
                    d["cur_l"] += 1
                    d["cur_w"] = 0
                    d["streak_l"] = max(d["streak_l"], d["cur_l"])
                    if d == t["all"]:
                        opp_ov = teams.get(si(match[f"team_{other}_id"]), {}).get("temp_value", 50.0)
                        d["opp_ov_l"].append(opp_ov)
                    if og - my_g > d["worst_loss"][0]: d["worst_loss"] = (og - my_g, teams.get(si(match[f"team_{other}_id"]), {}).get("name", "Desconhecido"))
    rows = []
    for tid, t in teams.items():
        if not t["all"]["games"]: continue
        all_d, bef, aft = t["all"], t["before"], t["after"]
        ov_init = sf(t.get("temp_value"))
        ov_aug = sf(t.get("wc_prev_edition_value")) or ov_init
        ov_final = sf(t.get("temp_value"))
        rows.append({
            "name": t["name"], "league": t["league"], "continent": t.get("continent_champ_name") or "Nenhuma", "players_count": len(t["players"]),
            "avg_age": sum(sf(p.get("age")) for p in t["players"]) / len(t["players"]) if t["players"] else 0,
            "all": all_d, "before": bef, "after": aft, "yellow_cards": t["yellow_cards"], "red_cards": t["red_cards"],
            "ov_init": ov_init, "ov_aug": ov_aug, "ov_final": ov_final, "value": si(t.get("temp_mk_value")), "salary": si(t.get("salary")), "bank": si(t.get("bank_balance")),
        })
    return sorted(rows, key=lambda r: -r["ov_final"])


def extract_data(save_path: Path):
    with tempfile.TemporaryDirectory() as td:
        with zipfile.ZipFile(save_path) as z:
            name = next(n for n in z.namelist() if re.fullmatch(r"save_.*\.db", Path(n).name))
            db_path = Path(z.extract(name, td))
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        emit(10, "Lendo configurações do save")
        date_info = get_date_info(cur)
        current_day = date_info["current_day"]
        managed_id = date_info["managed_id"]
        managed_name = date_info["managed_name"]
        
        leagues = {row["id"]: row["name"] for row in cur.execute("SELECT id,name FROM leagues2").fetchall()}
        teams = {row["id"]: row["name"] for row in cur.execute("SELECT id,name FROM teams2").fetchall()}
        emit(25, "Lendo jogadores")
        players = {}
        for row in cur.execute("SELECT * FROM players2").fetchall():
            player = dict(row)
            player["team_name"] = teams.get(si(player.get("team_id")), "Sem clube")
            player["league_name"] = leagues.get(si(player.get("champ_id")), "Sem liga")
            player["role"] = player.get("role") or "N/A"
            player["group"] = "GK" if player["role"] == "GK" else "DEF" if player["role"] in ("LB", "RB", "CB", "LWB", "RWB", "CD") else "MID" if player["role"] in ("CM", "LM", "RM", "CDM", "CAM") else "FWD"
            player["quality"] = sf(player.get("GKP")) if player["role"] == "GK" else sf(player.get("temp_value"))
            player["overall"] = player["quality"]
            
            # Correção: Campos s_ e h_ são blocos de texto (ex: CH_7:1 ou 0,0,0)
            player["games"] = parse_count_blocks(player.get("s_matches"))
            player["goals"] = parse_count_blocks(player.get("s_goals"))
            player["assists"] = parse_count_blocks(player.get("s_assists"))
            player["raw_clean"] = parse_count_blocks(player.get("s_clean_sheets"))
            player["raw_conceded"] = parse_count_blocks(player.get("s_g_conceded"))
            
            # Histórico de avaliações (bloco ex: 7.2|6.8)
            ratings_list = parse_rating_blocks(player.get("h_rating"))
            player["raw_rating"] = sum(ratings_list) / len(ratings_list) if ratings_list else 0.0
            player["rating_count"] = len(ratings_list)
            
            levels = parse_role_levels(player.get("roles_level"))
            player["role_scores"] = {role: round(float(levels.get(role, 0.0)), 1) for role in OUTFIELD_ROLES}
            player["season_overall_history"] = parse_values(player.get("s_value_in_time"))
            player["career_overall_history"] = parse_values(player.get("value_in_time"))
            player["yellow_cards"] = parse_count_blocks(player.get("s_y_cards"))
            player["red_cards"] = parse_count_blocks(player.get("s_r_cards"))
            player["market_value"] = si(player.get("temp_mk_value_modded")) or si(player.get("temp_mk_value"))
            player["salary"] = si(player.get("salary"))
            
            # Correção de Idade: birth_date é dias desde o início? Não, costuma ser dias relativos.
            # No save auditado: birth_date ~ -8707. 8707 / 365.25 = 23.8 anos.
            # Idade = (current_day - birth_date) / 365.25
            bday = sf(player.get("birth_date"))
            player["age"] = round((current_day - bday) / 365.25, 1)
            
            exp = si(player.get("contract_expiration"))
            player["contract_years"] = max(0.0, (exp - current_day) / 365.0) if exp else None
            player.update({
                "mvp": 0, "worst": 0, "rating_matches": 0, "rating_sum": 0.0, "apps": 0, "minutes": 0.0, 
                "clean_sheets": 0.0, "gk_apps": 0, "gk_minutes": 0.0, "gk_clean": 0.0, "gk_xg": 0.0, "gk_goals": 0.0,
                "pen_taken": 0, "pen_scored": 0, "pen_missed_total": 0, "pen_missed_out": 0, "pen_missed_saved": 0, "xgp_sum": 0.0,
                "gk_pen_faced": 0, "gk_pen_saved": 0, "gk_xgp_sum": 0.0
            })
            players[si(player["id"])] = player
        emit(38, "Calculando minutos, ratings e xG")
        all_matches = cur.execute("SELECT * FROM matches2 WHERE date<=? ORDER BY date,id", (current_day,)).fetchall()
        played_matches = [match for match in all_matches if si(match["state"], -1) == 2 and match["goals_1"] is not None and match["goals_2"] is not None]
        team_cards = defaultdict(lambda: {"yellow": 0, "red": 0})
        for raw in played_matches:
            match = dict(raw)
            for side in (1, 2):
                team_id = si(match.get(f"team_{side}_id"), -1)
                if team_id > 0:
                    team_cards[team_id]["yellow"] += max(0, si(match.get(f"yellows_{side}")))
                    team_cards[team_id]["red"] += max(0, si(match.get(f"reds_{side}")))
        club_summary = {"matches": 0, "shots": 0, "goals": 0, "xga": 0.0, "ga": 0}
        for raw in played_matches:
            match = dict(raw)
            sides = []
            for side in (1, 2):
                other = 3 - side
                lineup, entered, subbed = parse_ids(match.get(f"lineup_{side}")), parse_ids(match.get(f"entered_{side}")), parse_ids(match.get(f"subbed_{side}"))
                events = parse_events(match.get(f"events_{side}"))
                other_events = parse_events(match.get(f"events_{other}"))
                minutes = estimate_minutes(lineup, entered, subbed, events)
                ratings = parse_match_ratings(match.get(f"ratings_{side}"))
                if ratings:
                    high, low = max(ratings.values()), min(ratings.values())
                    for player_id, rating in ratings.items():
                        if player_id in players:
                            players[player_id]["rating_matches"] += 1
                            players[player_id]["rating_sum"] += rating
                            players[player_id]["mvp"] += rating == high
                            players[player_id]["worst"] += rating == low
                sides.append({"side": side, "team": si(match.get(f"team_{side}_id")), "minutes": minutes, "ratings": ratings, "goals": si(match.get(f"goals_{side}")), "xg": sf(match.get(f"x_goals_{side}")), "events": events, "other_events": other_events})
            
            # Pênaltis
            for side_idx, side in enumerate(sides):
                other = sides[1 - side_idx]
                # Códigos 7/8 = pênalti convertido/perdido; inclui desempate (minuto > 119)
                game_pens = [e for e in side["events"] if e[1] in (7, 8)]
                if not game_pens: continue

                other_side_id = other["side"]
                other_lineup = id_set(match.get(f"lineup_{other_side_id}"))
                other_events = other.get("events", [])

                def get_active_gk(minute):
                    # Identifica quem era o goleiro em campo no minuto dado
                    # 1. Quem começou o jogo como GK?
                    initial_gk = next((pid for pid in other_lineup if pid in players and players[pid]["role"] == "GK"), None)
                    # 2. Houve substituição envolvendo GK antes/nesse minuto?
                    # Código 6 = sai, Código 5 = entra
                    current_gk = initial_gk
                    gk_subs = sorted([e for e in other_events if e[1] in (5, 6) and e[2] in players and players[e[2]]["role"] == "GK"])
                    for m, code, pid in gk_subs:
                        if m > minute: break
                        if code == 6: current_gk = None
                        elif code == 5: current_gk = pid
                    # 3. O goleiro atual foi expulso antes/nesse minuto?
                    # Código 3 = vermelho direto, Código 4 = segundo amarelo
                    if current_gk:
                        reds = [e for e in other_events if e[1] in (3, 4) and e[2] == current_gk and e[0] <= minute]
                        if reds: current_gk = None
                    return current_gk

                for minute, code, pid in game_pens:
                    if pid not in players: continue
                    gk_id = get_active_gk(minute)
                    
                    b_ov = players[pid].get("SHO", 50.0)
                    g_ov = players[gk_id].get("GKP", 50.0) if gk_id else 50.0
                    diff = g_ov - b_ov
                    xgp = 0.750 - (diff * 0.015)
                    xgp = max(0.01, min(0.99, xgp))
                    
                    players[pid]["pen_taken"] += 1
                    saved = any(e[1] == 9 and e[0] == minute for e in other.get("events", []))
                    if code == 7:
                        players[pid]["pen_scored"] += 1
                        # Reconciliar: se o save não contou o gol (bug do save), nós contamos
                        # E garantimos que 'gols sem pênalti' não fique negativo
                        players[pid]["goals"] = max(players[pid]["goals"], players[pid]["pen_scored"])
                    else:
                        players[pid]["pen_missed_total"] += 1
                        if saved:
                            players[pid]["pen_missed_saved"] += 1
                        else:
                            players[pid]["pen_missed_out"] += 1
                    players[pid]["xgp_sum"] += xgp
                    
                    if gk_id:
                        if code == 7 or saved:
                            players[gk_id]["gk_pen_faced"] += 1
                            players[gk_id]["gk_xgp_sum"] += xgp
                            if saved:
                                players[gk_id]["gk_pen_saved"] += 1

            for side in sides:
                other = sides[1] if side["side"] == 1 else sides[0]
                player_ids = set(side["minutes"]) | set(side["ratings"])
                total_gk_minutes = sum(side["minutes"].get(player_id, 0.0) for player_id in player_ids if player_id in players and players[player_id]["role"] == "GK")
                for player_id in player_ids:
                    if player_id not in players:
                        continue
                    minutes = side["minutes"].get(player_id, 45.0 if player_id in side["ratings"] else 0.0)
                    if minutes <= 0:
                        continue
                    player = players[player_id]
                    player["apps"] += 1
                    player["minutes"] += minutes
                    if other["goals"] == 0 and player["group"] == "DEF":
                        player["clean_sheets"] += 1
                    if player["role"] == "GK":
                        player["gk_apps"] += 1
                        player["gk_minutes"] += minutes
                        player["gk_clean"] += other["goals"] == 0
                        share = minutes / total_gk_minutes if total_gk_minutes else 1.0
                        player["gk_xg"] += other["xg"] * share
                        player["gk_goals"] += other["goals"] * share
            if si(match.get("team_1_id")) == managed_id or si(match.get("team_2_id")) == managed_id:
                side = 1 if si(match.get("team_1_id")) == managed_id else 2
                other = 3 - side
                club_summary["matches"] += 1
                club_summary["shots"] += si(match.get(f"shots_on_{side}")) + si(match.get(f"shots_off_{side}"))
                club_summary["goals"] += si(match.get(f"goals_{side}"))
                club_summary["xga"] += sf(match.get(f"x_goals_{other}"))
                club_summary["ga"] += si(match.get(f"goals_{other}"))
        emit(56, "Calculando overall Footlord")
        for player in players.values():
            if player["minutes"] <= 0 and player["games"]:
                player["minutes"], player["apps"] = player["games"] * 45.0, int(player["games"])
            if player["rating_matches"] == 0 and player["rating_count"]:
                player["rating_matches"], player["rating_sum"] = player["rating_count"], player["raw_rating"] * player["rating_count"]
            player["rating"] = player["rating_sum"] / player["rating_matches"] if player["rating_matches"] else player["raw_rating"]
            player["ga"] = player["goals"] + player["assists"]
            player["goals_pg"] = player["goals"] / player["games"] if player["games"] else 0.0
            player["assists_pg"] = player["assists"] / player["games"] if player["games"] else 0.0
            player["ga_pg"] = player["ga"] / player["games"] if player["games"] else 0.0
            player["minutes_goal"] = player["minutes"] / player["goals"] if player["goals"] else None
            player["minutes_assist"] = player["minutes"] / player["assists"] if player["assists"] else None
            player["mvp_pct"] = player["mvp"] / player["rating_matches"] if player["rating_matches"] else 0.0
            player["worst_pct"] = player["worst"] / player["rating_matches"] if player["rating_matches"] else 0.0
            if player["group"] == "DEF":
                player["clean_sheets"] = max(player["clean_sheets"], player["raw_clean"])
            if player["role"] == "GK" and not player["gk_apps"] and player["games"]:
                player["gk_apps"], player["gk_minutes"], player["gk_clean"], player["gk_goals"] = int(player["games"]), player["minutes"], player["raw_clean"], player["raw_conceded"]
            if player["role"] == "GK":
                player["output"] = (player["gk_xg"] - player["gk_goals"]) / (player["gk_minutes"] / 90) if player["gk_minutes"] else 0.0
            elif player["group"] == "DEF":
                player["output"] = 0.65 * (player["clean_sheets"] / player["games"] if player["games"] else 0) + 0.35 * min(1, player["ga_pg"])
            else:
                player["output"] = min(1, player["ga_pg"])
            player["rating_score"] = clip((player["rating"] - 5.0) / 3 * 100) if player["rating"] else 0
            player["efficiency"] = player["quality"] / (1 + max(0, player["market_value"]) / 1_000_000 + 0.5 * max(0, player["salary"]) / 1_000_000)
        group_outputs = defaultdict(list)
        for player in players.values():
            if player["games"] or player["role"] == "GK":
                group_outputs[player["group"]].append(player["output"])
        efficiencies = [player["efficiency"] for player in players.values()]
        for player in players.values():
            output_score, cost_score = percentile(group_outputs[player["group"]], player["output"]), percentile(efficiencies, player["efficiency"])
            age_val = sf(player.get("age"), 34)
            age_score = clip((34 - age_val) / 16 * 100)
            player["score"] = clip(0.50 * player["quality"] + 0.20 * player["rating_score"] + 0.15 * output_score + 0.10 * cost_score + 0.05 * age_score)
        ordered_players = sorted(players.values(), key=lambda player: (-player["overall"], player["market_value"], player["id"]))
        goalkeepers = sorted([player for player in players.values() if player["role"] == "GK" and player["gk_apps"]], key=lambda player: (-player["overall"], -player["gk_apps"], player["id"]))
        team_rows = build_team_rows(cur, date_info, players, played_matches, team_cards)
        conn.close()
        return {"current_day": current_day, "club": managed_name, "players": ordered_players, "gks": goalkeepers, "teams": team_rows, "club_summary": club_summary, "played_matches": len(played_matches)}


def player_values(players):
    rows = []
    for rank, player in enumerate(players, 1):
        name = " ".join(part for part in (str(player.get("name") or "").strip(), str(player.get("surname") or "").strip()) if part) or f"ID {player['id']}"
        current = round(player["overall"], 1)
        season_start = player["season_overall_history"][0] if player["season_overall_history"] else current
        career_start = player["career_overall_history"][0] if player["career_overall_history"] else season_start
        # Usar apps (reconstruído das partidas) se games (bruto do save) estiver zerado
        games_val = player["apps"] if player["apps"] > 0 else player["games"]
        games, yellow, red = round(games_val), round(player["yellow_cards"]), round(player["red_cards"])
        
        pen_taken = player["pen_taken"]
        pen_scored = player["pen_scored"]
        pen_missed_total = player["pen_missed_total"]
        pen_missed_out = player["pen_missed_out"]
        xgp = player["xgp_sum"]
        pen_rate = pen_scored / pen_taken if pen_taken else None
        pen_vs_xgp = pen_scored - xgp if pen_taken else None
        
        role_cols = [player.get("role_scores", {}).get(role, 0.0) for role in OUTFIELD_ROLES]
        # Decimais no máximo milésimo (3 casas)
        def r3(v):
            return None if v is None else round(float(v), 3)
        def r1(v):
            return None if v is None else round(float(v), 1)
        rows.append([
            rank, name, player["role"], *role_cols, player["age"], player.get("team_name") or "Sem clube", player.get("league_name") or "Sem liga", player.get("league_division"), player.get("nationality") or player.get("nation_name") or "",
            r1(current), r1(season_start), r1(current - season_start), r1(career_start), r1(current - career_start), games, yellow, red, yellow + red, r3(yellow / games) if games else None, r3(red / games) if games else None,
            round(player["goals"]), round(player["goals"] - pen_scored), pen_taken, pen_scored, pen_missed_total, pen_missed_out, r3(pen_rate), r3(xgp), r3(pen_vs_xgp),
            round(player["assists"]), round(player["ga"]), r3(player["goals"] / games) if games else 0.0, r3(player["assists"] / games) if games else 0.0, r3(player["ga"] / games) if games else 0.0, r1(player["rating"]) if player["rating"] else None,
            round(player["minutes"]), r1(player["minutes"] / player["apps"]) if player["apps"] else None, r1(player["minutes_goal"]) if player.get("minutes_goal") else None, r1(player["minutes_assist"]) if player.get("minutes_assist") else None,
            player["mvp"], player["worst"], r3(player["mvp_pct"]), r3(player["worst_pct"]), round(player["clean_sheets"]) if player["group"] == "DEF" else None,
            player["market_value"], player["salary"], r1(player["contract_years"]) if player["contract_years"] is not None else None, "Sem clube" if si(player.get("team_id"), -1) <= 0 else ("Em empréstimo" if si(player.get("loan_status")) else "Contrato"),
            "Sim" if si(player.get("loan_status")) else "Não", "Sim" if si(player.get("transfer_status")) == 1 else "Não", r1(sf(player.get("PAC"))), r1(sf(player.get("SHO"))), r1(sf(player.get("PAS"))), r1(sf(player.get("DEF"))), r1(sf(player.get("PHY"))), r1(sf(player.get("MEN"))), r1(sf(player.get("GKP"))), player["id"],
        ])
    return rows


def goalkeeper_values(goalkeepers):
    rows = []
    for rank, player in enumerate(goalkeepers, 1):
        xg, conceded, apps, minutes = player["gk_xg"], player["gk_goals"], player["gk_apps"], player["gk_minutes"]
        prevented = xg - conceded
        rate = prevented / xg if xg else None
        
        pen_faced = player["gk_pen_faced"]
        pen_saved = player["gk_pen_saved"]
        xgp = player["gk_xgp_sum"]
        pen_save_rate = pen_saved / pen_faced if pen_faced else None
        pen_vs_xgp = pen_saved - (pen_faced - xgp) if pen_faced else None # xGP é gol esperado, logo defesa esperada é faced - xGP
        
        classification = "Amostra curta" if apps < 5 else "Sem xG" if rate is None else "Excelente" if rate >= .15 else "Bom" if rate >= .05 else "Neutro" if rate >= -.05 else "Crítico"
        name = " ".join(part for part in (str(player.get("name") or "").strip(), str(player.get("surname") or "").strip()) if part) or f"ID {player['id']}"
        def r3(v):
            return None if v is None else round(float(v), 3)
        def r1(v):
            return None if v is None else round(float(v), 1)
        rows.append([
            rank, name, player.get("team_name") or "Sem clube", player.get("league_name") or "Sem liga", apps, round(minutes), round(player["gk_clean"]), r3(player["gk_clean"] / apps) if apps else None, r3(xg), r3(conceded), r3(conceded / apps) if apps else None, r3(prevented), r3(prevented / apps) if apps else None, r3(prevented / (minutes / 90)) if minutes else None, r3(rate), r3(conceded / xg) if xg else None,
            pen_faced, pen_saved, r3(pen_save_rate), r3(xgp), r3(pen_vs_xgp),
            classification, "Sim" if apps >= 5 else "Não", r1(player["overall"]), player["id"]
        ])
    return rows


def percentiles_for_rows(rows, column_indices):
    result = {}
    for column in column_indices:
        values = sorted(float(row[column]) for row in rows if column < len(row) and isinstance(row[column], (int, float)) and math.isfinite(float(row[column])))
        result[column] = values
    return result


def color_name(value, values, inverse=False):
    if value is None or not isinstance(value, (int, float)) or not values:
        return "neutral"
    lower, medium, high = values[max(0, int((len(values) - 1) * .20))], values[max(0, int((len(values) - 1) * .50))], values[max(0, int((len(values) - 1) * .80))]
    if inverse:
        return "excellent" if value <= lower else "good" if value <= medium else "medium" if value <= high else "bad"
    return "bad" if value <= lower else "medium" if value <= medium else "good" if value <= high else "excellent"


def set_cell_style(cell, fill="neutral", number_format=None, alignment=CENTER, font=DATA_FONT):
    cell.font, cell.fill, cell.border, cell.alignment = font, FILLS[fill], BORDER, alignment
    if number_format:
        cell.number_format = number_format

def write_sheet(wb, title, subtitle, headers, rows, number_formats, color_columns=None, inverse_columns=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    # Título e Subtítulo
    ws.append([title.upper()])
    set_cell_style(ws.cell(row=1, column=1), fill="title", font=TITLE_FONT, alignment=LEFT_WRAP)
    ws.append([subtitle])
    set_cell_style(ws.cell(row=2, column=1), fill="subtitle", font=HEADER_FONT, alignment=LEFT_WRAP)
    
    # Legenda
    legend = ["LEGENDA:", "RUIM", "MÉDIO", "BOM", "EXCELENTE"]
    legend_fills = ["neutral", "bad", "medium", "good", "excellent"]
    ws.append(legend)
    for i, f in enumerate(legend_fills, 1):
        set_cell_style(ws.cell(row=3, column=i), fill=f)
    
    # Cabeçalhos (wrap só no header)
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        set_cell_style(ws.cell(row=4, column=i), fill="header", font=HEADER_FONT, alignment=CENTER_WRAP)
        
    # Larguras — fileiras finas: colunas Role estreitas, dados sem wrap
    for i, h in enumerate(headers, 1):
        h_str = str(h or "").strip()
        h_low = h_str.casefold()
        if h_str.startswith("Role ") or h_str.startswith("ROLE "):
            w = 7
        elif any(x in h_str for x in ("Jogador", "Player", "Goleiro", "Goalkeeper")):
            w = 22
        elif any(x in h_str for x in ("Time", "Team")):
            w = 18
        elif any(x in h_str for x in ("Liga", "League", "Adversário", "Opponent", "Para quem")):
            w = 18
        elif "€" in h_str:
            w = 14
        elif any(x in h_str for x in ("Overall", "overall", "xGP")):
            w = 12
        else:
            w = 11
        ws.column_dimensions[get_column_letter(i)].width = w

    # Dados (wrap_text=False via CENTER → fileiras finas)
    percentiles = percentiles_for_rows(rows, color_columns or [])
    for row_idx, row_data in enumerate(rows, 5):
        ws.append(row_data)
        for col_idx, val in enumerate(row_data, 1):
            fill = "neutral"
            if color_columns and (col_idx - 1) in color_columns:
                fill = color_name(val, percentiles.get(col_idx - 1), inverse=(inverse_columns and (col_idx - 1) in inverse_columns))
            set_cell_style(ws.cell(row=row_idx, column=col_idx), fill=fill, number_format=number_formats.get(col_idx - 1), alignment=CENTER)
            
    last = max(4, 4 + len(rows))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last}"


def create_workbook(data, output_path, language="pt"):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Players 2
    # Índices após expandir Role CD..ST (11 colunas no lugar de 1):
    # 0 Rank, 1 Nome, 2 Pos, 3-13 Roles, 14 Idade, 15 Time, 16 Liga, 17 Div, 18 País,
    # 19 Overall, 20 ov_temp, 21 var_temp, 22 ov_car, 23 var_car,
    # 24 jogos, 25 Y, 26 R, 27 tot, 28 Y/p, 29 R/p,
    # 30 gols, 31 gols_sem_pen, 32 pen_bat, 33 gols_pen, 34 pen_perd, 35 pen_fora, 36 taxa_pen, 37 xgp, 38 pen_vs_xgp,
    # 39 ast, 40 GA, 41 G/p, 42 A/p, 43 GA/p, 44 média, 45 min,
    # 46 min/p, 47 min/gol, 48 min/ast, 49 MVP, 50 pior, 51 MVP%, 52 pior%, 53 clean,
    # 54 mk, 55 sal, 56 contr, 57 status, 58 loan, 59 list, 60-66 attrs, 67 id
    p_rows = player_values(data["players"])
    p_fmts = {
        **{i: NUM_FORMAT_DEC1 for i in range(3, 14)},  # roles
        14: '0',
        19: NUM_FORMAT_DEC1, 20: NUM_FORMAT_DEC1, 21: NUM_FORMAT_DEC1, 22: NUM_FORMAT_DEC1, 23: NUM_FORMAT_DEC1,
        24: '0', 25: '0', 26: '0', 27: '0', 28: NUM_FORMAT_DEC3, 29: NUM_FORMAT_DEC3,
        30: '0', 31: '0', 32: '0', 33: '0', 34: '0', 35: '0', 36: NUM_FORMAT_PCT, 37: NUM_FORMAT_DEC3, 38: NUM_FORMAT_DEC3,
        39: '0', 40: '0', 41: NUM_FORMAT_DEC3, 42: NUM_FORMAT_DEC3, 43: NUM_FORMAT_DEC3, 44: NUM_FORMAT_DEC1,
        45: NUM_FORMAT_INT, 46: NUM_FORMAT_DEC1, 47: NUM_FORMAT_DEC1, 48: NUM_FORMAT_DEC1,
        49: '0', 50: '0', 51: NUM_FORMAT_PCT, 52: NUM_FORMAT_PCT, 53: '0',
        54: NUM_FORMAT_INT, 55: NUM_FORMAT_INT, 56: NUM_FORMAT_DEC1,
        **{i: NUM_FORMAT_DEC1 for i in range(60, 67)}, 67: '0',
    }
    # Colunas de comparação coloridas (regra percentil já usada)
    p_colors = list(range(3, 14)) + [14, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56] + list(range(60, 67))
    # Menor é melhor: idade, cartões, min/gol, min/ast, pior, pior%, salário
    p_inv = [14, 25, 26, 27, 28, 29, 47, 48, 50, 52, 55]
    write_sheet(wb, "Players 2", f"Relatório de Jogadores - {data['club']} ({data['current_day']} dias)", localized_headers(PLAYERS_HEADERS, language), p_rows, p_fmts, p_colors, p_inv)
    
    # Goleiros xG
    g_rows = goalkeeper_values(data["gks"])
    g_fmts = {
        5: NUM_FORMAT_INT, 6: '0', 7: NUM_FORMAT_DEC3, 8: NUM_FORMAT_DEC3, 9: NUM_FORMAT_DEC3, 10: NUM_FORMAT_DEC3,
        11: NUM_FORMAT_DEC3, 12: NUM_FORMAT_DEC3, 13: NUM_FORMAT_DEC3, 14: NUM_FORMAT_PCT, 15: NUM_FORMAT_DEC3,
        16: '0', 17: '0', 18: NUM_FORMAT_PCT, 19: NUM_FORMAT_DEC3, 20: NUM_FORMAT_DEC3, 23: NUM_FORMAT_DEC1, 24: '0',
    }
    # Comparação: clean/p, evitados, taxa, gols/xG, taxa defesa pen, vs xGP, overall
    g_colors = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 23]
    g_inv = [9, 10, 15]  # gols sofridos, gols/partida, gols/xG (menor melhor)
    write_sheet(wb, "Goleiros xG", f"Desempenho de Goleiros - {data['club']}", localized_headers(GK_HEADERS, language), g_rows, g_fmts, g_colors, g_inv)
    
    # Times
    t_rows = []
    for rank, t in enumerate(data["teams"], 1):
        all_d, bef, aft = t["all"], t["before"], t["after"]
        t_rows.append([
            rank, t["name"], t["league"], t["continent"], t["players_count"], round(t["avg_age"], 1),
            all_d["games"], all_d["wins"], all_d["draws"], all_d["losses"], t["yellow_cards"], t["red_cards"], t["yellow_cards"] + t["red_cards"], round(t["yellow_cards"] / all_d["games"], 3) if all_d["games"] else None, round(t["red_cards"] / all_d["games"], 3) if all_d["games"] else None,
            bef["games"], bef["wins"], bef["draws"], bef["losses"],
            aft["games"], aft["wins"], aft["draws"], aft["losses"],
            bef["streak_w"], bef["streak_l"], aft["streak_w"], aft["streak_l"],
            all_d["goals_for"], all_d["goals_against"], round(all_d["goals_for"] / all_d["games"], 3) if all_d["games"] else None,
            round(all_d["goals_against"] / all_d["games"], 3) if all_d["games"] else None,
            all_d["shots"], round(all_d["shots"] / all_d["games"], 3) if all_d["games"] else None, all_d["on_target"], round(all_d["on_target"] / all_d["shots"], 3) if all_d["shots"] else None, round(all_d["goals_for"] / all_d["on_target"], 3) if all_d["on_target"] else None,
            bef["goals_for"], bef["goals_against"], round(bef["goals_for"] / bef["games"], 3) if bef["games"] else None, round(bef["goals_against"] / bef["games"], 3) if bef["games"] else None,
            aft["goals_for"], aft["goals_against"], round(aft["goals_for"] / aft["games"], 3) if aft["games"] else None, round(aft["goals_against"] / aft["games"], 3) if aft["games"] else None,
            all_d["failed_to_score"], all_d["clean_sheets"], bef["failed_to_score"], bef["clean_sheets"], aft["failed_to_score"], aft["clean_sheets"],
            all_d["biggest_win"][0], all_d["biggest_win"][1], all_d["worst_loss"][0], all_d["worst_loss"][1],
            bef["biggest_win"][0], bef["biggest_win"][1], bef["worst_loss"][0], bef["worst_loss"][1],
            aft["biggest_win"][0], aft["biggest_win"][1], aft["worst_loss"][0], aft["worst_loss"][1],
            t["ov_init"], t["ov_aug"], t["ov_final"], None, None, None,
            sum(all_d["opp_ov_w"]) / len(all_d["opp_ov_w"]) if all_d["opp_ov_w"] else None,
            sum(all_d["opp_ov_d"]) / len(all_d["opp_ov_d"]) if all_d["opp_ov_d"] else None,
            sum(all_d["opp_ov_l"]) / len(all_d["opp_ov_l"]) if all_d["opp_ov_l"] else None,
            t["value"], t["salary"], t["bank"], None, None, None, None, None, None, None, None,
        ])
    t_fmts = {
        5: NUM_FORMAT_DEC1, 13: NUM_FORMAT_DEC3, 14: NUM_FORMAT_DEC3,
        29: NUM_FORMAT_DEC3, 30: NUM_FORMAT_DEC3, 32: NUM_FORMAT_DEC3, 34: NUM_FORMAT_PCT, 35: NUM_FORMAT_PCT,
        38: NUM_FORMAT_DEC3, 39: NUM_FORMAT_DEC3, 42: NUM_FORMAT_DEC3, 43: NUM_FORMAT_DEC3,
        63: NUM_FORMAT_DEC1, 64: NUM_FORMAT_DEC1, 65: NUM_FORMAT_DEC1,
        69: NUM_FORMAT_DEC1, 70: NUM_FORMAT_DEC1, 71: NUM_FORMAT_DEC1,
        72: NUM_FORMAT_INT, 73: NUM_FORMAT_INT, 74: NUM_FORMAT_INT,
    }
    # Comparativos de desempenho/disciplina/eficiência
    t_colors = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,
                27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49,
                63, 64, 65, 69, 70, 71, 72, 73, 74]
    t_inv = [9, 11, 12, 13, 14, 18, 22, 24, 26, 28, 30, 37, 39, 41, 43, 44, 46, 48]
    write_sheet(wb, "Times", f"Relatório de Equipes - {data['club']}", localized_headers(TIMES_HEADERS, language), t_rows, t_fmts, t_colors, t_inv)
    
    # Resumo e Dicionário
    write_sheet(wb, "Resumo", "Resumo do Processamento", ["Indicador", "Valor"], [["Clube Gerenciado", data["club"]], ["Dia Interno", data["current_day"]], ["Jogadores Processados", len(data["players"])], ["Goleiros Elegíveis", len(data["gks"])], ["Times Analisados", len(data["teams"])], ["Partidas Concluídas", data["played_matches"]]], {})
    write_sheet(wb, "Dicionário", "Significado das Colunas", ["Campo", "Leitura"], [
        ["xGP", "Gols esperados de pênalti (base 0,75 ajustado por overall SHO vs GKP)"],
        ["Gols pênalti vs xGP", "Diferença entre gols reais e esperados (positivo = acima da média)"],
        ["Taxa conversão", "Percentual de acerto nas cobranças (inclui pênaltis de desempate, minuto > 119)"],
        ["Pênaltis (desempate)", "Eventos com código 7/8 após o 119' entram nas mesmas colunas de pênalti"],
        ["Role CD…ST", "Nota 0–10 daquele papel no jogador (0 se não tiver); GK não tem coluna própria aqui"],
        ["Cores", "Percentis na coluna: vermelho ruim → azul excelente (colunas invertidas: menor é melhor)"],
    ], {})
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--save")
    parser.add_argument("--template")
    parser.add_argument("--output", required=True)
    parser.add_argument("--language", default="pt")
    parser.add_argument("--template-only", action="store_true")
    args = parser.parse_args()
    try:
        if args.template_only:
            # Gerar apenas o modelo limpo (headers e legendas)
            data = {
                "current_day": 0, "club": "Modelo", "players": [], "gks": [], "teams": [],
                "club_summary": {"matches": 0, "shots": 0, "goals": 0, "xga": 0.0, "ga": 0},
                "played_matches": 0
            }
            create_workbook(data, args.output, args.language)
        else:
            if not args.save:
                print("ERRO: --save é obrigatório quando não usar --template-only", file=sys.stderr)
                sys.exit(1)
            data = extract_data(Path(args.save))
            create_workbook(data, args.output, args.language)
            emit(100, "Relatório concluído com sucesso")
    except Exception as e:
        print(f"ERRO: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
